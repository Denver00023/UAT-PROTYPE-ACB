import streamlit as st
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import shutil



# CONFIGURATION

DATABASE_FILE = Path("database.xlsx")
BACKUP_FOLDER = Path("backup")


# LOAD EXCEL

@st.cache_data
def load_excel(file_path, file_modified_time):

    return pd.read_excel(
        file_path,
        sheet_name=None,
        dtype=str
    )



# SEARCH FUNCTION

def search_data(df, search_text):

    if not search_text.strip():
        return df

    search_text = search_text.strip().lower()

    searchable = df.fillna("").astype(str)

    mask = searchable.apply(
        lambda column: column.str.lower().str.contains(
            search_text,
            na=False,
            regex=False
        )
    ).any(axis=1)

    return df[mask]



# BACKUP EXCEL FILE

def create_backup():

    BACKUP_FOLDER.mkdir(
        exist_ok=True
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    backup_file = (
        BACKUP_FOLDER
        / f"database_backup_{timestamp}.xlsx"
    )

    shutil.copy2(
        DATABASE_FILE,
        backup_file
    )

    return backup_file


# SAVE UPDATED DATAFRAME TO ONE SHEET

def save_sheet(sheet_name, df):


    # CREATE BACKUP BEFORE MODIFYING DATABASE

    backup_file = create_backup()

    try:
        
        # LOAD EXISTING WORKBOOK
        
        workbook = load_workbook(
            DATABASE_FILE
        )

        # CHECK SHEET
        
        if sheet_name not in workbook.sheetnames:

            raise ValueError(
                f"Sheet '{sheet_name}' does not exist."
            )

        worksheet = workbook[sheet_name]

        
        # DELETE EXISTING DATA
        
        if worksheet.max_row > 0:

            worksheet.delete_rows(
                1,
                worksheet.max_row
            )

        
        # WRITE COLUMN HEADERS
        
        for column_number, column_name in enumerate(
            df.columns,
            start=1
        ):

            worksheet.cell(
                row=1,
                column=column_number,
                value=column_name
            )

        
        # WRITE DATA

        for row_number, row in enumerate(
            df.itertuples(index=False),
            start=2
        ):

            for column_number, value in enumerate(
                row,
                start=1
            ):

                if pd.isna(value):

                    value = ""

                worksheet.cell(
                    row=row_number,
                    column=column_number,
                    value=str(value)
                )

        # SAVE DATABASE
        workbook.save(
            DATABASE_FILE
        )

        # CLEAR CACHE
        st.cache_data.clear()

        return backup_file

    except PermissionError:

        raise PermissionError(
            "database.xlsx is locked or you do not have "
            "permission to modify it. "
            "Please close the Excel file and try again."
        )

# SUCCESS POPUP / MODAL
def show_success_popup(title, message, backup_file=None):

    @st.dialog(title)
    def popup():

        st.success(message)

        if backup_file:

            st.info(
                f"📦 Backup created: {backup_file.name}"
            )

        st.write(
            "The database has been updated successfully."
        )

        if st.button(
            "OK",
            type="primary",
            use_container_width=True
        ):

            st.rerun()

    popup()

# MAIN APP

def run():

    st.set_page_config(
        page_title="Customs Reference Search",
        page_icon="🔎",
        layout="wide"
    )

    st.title(
        "🔎 Customs Reference Search Engine"
    )

    st.caption(
        "Search, add, edit and delete customs reference data."
    )

    
    # CHECK DATABASE
    
    if not DATABASE_FILE.exists():

        st.error(
            f"Excel file not found: "
            f"{DATABASE_FILE.resolve()}"
        )

        return

    # LOAD DATABASE
    
    try:

        sheets = load_excel(
            DATABASE_FILE,
            DATABASE_FILE.stat().st_mtime
        )

    except Exception as e:

        st.error(
            f"Excel reading error: {e}"
        )

        return

    
    # DATABASE STATUS
    
    st.success(
        f"✅ Database loaded successfully: "
        f"{len(sheets)} sheets"
    )

    # SHOW DATABASE SHEETS
    
    st.write(
        "### 📚 Database Sheets"
    )

    for number, sheet_name in enumerate(
        sheets.keys(),
        start=1
    ):

        df = sheets[sheet_name]

        st.write(
            f"**{number}. {sheet_name}** — "
            f"{len(df):,} rows × "
            f"{len(df.columns):,} columns"
        )

    st.divider()
    
    # SELECT SHEET
    
    sheet_names = list(
        sheets.keys()
    )

    selected_sheet = st.selectbox(
        "Select Database Sheet",
        sheet_names
    )

    df = sheets[
        selected_sheet
    ].copy()

    st.subheader(
        f"📋 Database: {selected_sheet}"
    )

    # ACTION
    
    action = st.radio(
        "Action",
        [
            "🔎 Search",
            "➕ Add Record",
            "✏️ Edit Record",
            "🗑️ Delete Record"
        ],
        horizontal=True
    )

    st.divider()

    # SEARCH
    
    if action == "🔎 Search":

        search_text = st.text_input(
            "🔎 Search / Find",
            placeholder=(
                "Type HS Code, product, airline, "
                "carrier, destination, code, etc."
            )
        )

        results = search_data(
            df,
            search_text
        )

        if search_text.strip():

            st.write(
                f"### 🔍 Found "
                f"{len(results):,} result(s)"
            )

        else:

            st.write(
                f"### 📊 Showing all "
                f"{len(results):,} records"
            )

        st.dataframe(
            results,
            use_container_width=True,
            hide_index=True
        )

    
    # ADD RECORD
    
    elif action == "➕ Add Record":

        st.subheader(
            "➕ Add New Record"
        )

        new_values = {}

        # Create input for every column
        for column in df.columns:

            new_values[column] = st.text_input(
                column,
                key=f"add_{selected_sheet}_{column}"
            )

        if st.button(
            "💾 Add Record",
            type="primary"
        ):

            new_row = pd.DataFrame(
                [new_values]
            )

            updated_df = pd.concat(
                [df, new_row],
                ignore_index=True
            )

            try:

                backup = save_sheet(
                    selected_sheet,
                    updated_df
                )

                show_success_popup(
                    "✅ Record Added",
                    f"Record successfully added to '{selected_sheet}'.",
                    backup
                )

            except Exception as e:

                st.error(
                    f"Unable to save record: {e}"
                )

    
    # EDIT RECORD

    elif action == "✏️ Edit Record":

        st.subheader(
            "✏️ Edit Existing Record"
        )

        if df.empty:

            st.warning(
                "This sheet has no records."
            )

            return

        # Select row
        row_number = st.number_input(
            "Row number to edit",
            min_value=1,
            max_value=len(df),
            value=1,
            step=1
        )

        row_index = int(
            row_number - 1
        )

        st.write(
            f"Editing Excel data row "
            f"**{row_number}**"
        )

        edited_values = {}

        for column in df.columns:

            current_value = str(
                df.iloc[
                    row_index
                ][column]
            )

            edited_values[column] = st.text_input(
                column,
                value=current_value,
                key=(
                    f"edit_"
                    f"{selected_sheet}_"
                    f"{row_index}_"
                    f"{column}"
                )
            )

        if st.button(
            "💾 Save Changes",
            type="primary"
        ):

            updated_df = df.copy()

            for column in df.columns:

                updated_df.loc[
                    row_index,
                    column
                ] = edited_values[column]

            try:

                backup = save_sheet(
                    selected_sheet,
                    updated_df
                )

                show_success_popup(
                    "✅ Record Updated",
                    f"Record successfully updated in '{selected_sheet}'.",
                    backup
                )

            except Exception as e:

                st.error(
                    f"Unable to save changes: {e}"
                )

    # DELETE RECORD

    elif action == "🗑️ Delete Record":

        st.subheader(
            "🗑️ Delete Record"
        )

        if df.empty:

            st.warning(
                "This sheet has no records."
            )

            return

        row_number = st.number_input(
            "Row number to delete",
            min_value=1,
            max_value=len(df),
            value=1,
            step=1
        )

        row_index = int(
            row_number - 1
        )

        st.write(
            "### Record to be deleted"
        )

        st.dataframe(
            df.iloc[
                [row_index]
            ],
            use_container_width=True,
            hide_index=True
        )

        confirm_delete = st.checkbox(
            "I confirm that I want to delete this record."
        )

        if st.button(
            "🗑️ Delete Record",
            type="primary",
            disabled=not confirm_delete
        ):

            updated_df = df.drop(
                index=row_index
            ).reset_index(
                drop=True
            )

            try:

                backup = save_sheet(
                    selected_sheet,
                    updated_df
                )

                show_success_popup(
                    "🗑️ Record Deleted",
                    f"Record successfully deleted from '{selected_sheet}'.",
                    backup
                )


            except Exception as e:

                st.error(
                    f"Unable to delete record: {e}"
                )

# RUN

if __name__ == "__main__":
    run()
